from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.excel_utils import read_cell_after_recalc, recalculate_and_save


def to_fraction(value: Any) -> float:
    """
    Converte valores como:
    - 0.23 -> 0.23
    - 23 -> 0.23
    - '23,00%' -> 0.23
    - '0,23%' -> 0.23
    """
    if value is None or value == "":
        return 0.0

    if isinstance(value, (int, float)):
        number = float(value)
        return number / 100 if number > 1 else number

    if isinstance(value, str):
        text = value.strip().replace("%", "").replace(" ", "").replace(",", ".")
        if not text:
            return 0.0
        number = float(text)
        return number / 100 if number > 1 else number

    raise ValueError(f"Não foi possível interpretar o valor: {value!r}")


def find_row_by_label(ws, label: str) -> int:
    """
    Procura um texto na primeira coluna e retorna a linha correspondente.
    """
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_value, str) and cell_value.strip() == label:
            return row_idx

    raise ValueError(f"Não encontrei a linha com o texto: {label}")


def find_next_blank_column_after_last_value(ws, row_idx: int, start_col: int = 2) -> int:
    """
    Encontra a próxima coluna vazia depois do último valor preenchido na linha.
    """
    last_filled_col = None

    for col_idx in range(start_col, ws.max_column + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value not in (None, ""):
            last_filled_col = col_idx

    if last_filled_col is None:
        return start_col

    return last_filled_col + 1


def update_curve_s(config: dict, dashboard_path: Path) -> Path:
    """
    Lê o % Real do Dashboard atualizado e escreve a Curva S.
    A coluna de atualização é definida pelo próximo espaço em branco
    após o último valor preenchido na linha Realizado Acumulado.
    """
    input_folder = Path(config["paths"]["input_folder"])
    output_folder = Path(config["paths"]["output_folder"])

    curve_file = config["paths"]["curva_s_file"]
    curve_input_path = input_folder / curve_file
    curve_output_path = output_folder / curve_file

    if not curve_input_path.exists():
        raise FileNotFoundError(f"Curva S não encontrada: {curve_input_path}")

    dashboard_sheet = config["dashboard"]["cronograma_sheet"]
    dashboard_cell = config["dashboard"]["cronograma_real_cell"]

    # Lê o % Real do dashboard já recalculado
    current_real_value = read_cell_after_recalc(dashboard_path, dashboard_sheet, dashboard_cell)
    current_real_percent = to_fraction(current_real_value)

    wb = load_workbook(curve_input_path)
    ws = wb[config["curve_s"]["sheet_name"]]

    label_acum = config["curve_s"]["realizado_acumulado_label"]
    label_week = config["curve_s"]["realizado_semanal_label"]

    row_acum = find_row_by_label(ws, label_acum)
    row_week = find_row_by_label(ws, label_week)

    # Próxima coluna vazia após o último acumulado
    target_col = find_next_blank_column_after_last_value(ws, row_acum, start_col=2)

    # Valor da semana anterior
    previous_col = target_col - 1
    previous_acum = 0.0

    if previous_col >= 2:
        prev_value = ws.cell(row=row_acum, column=previous_col).value
        previous_acum = to_fraction(prev_value)

    current_acum = current_real_percent
    current_week = current_acum - previous_acum

    # Escreve na Curva S
    cell_acum = ws.cell(row=row_acum, column=target_col)
    cell_acum.value = current_acum
    cell_acum.number_format = "0.00%"

    cell_week = ws.cell(row=row_week, column=target_col)
    cell_week.value = current_week
    cell_week.number_format = "0.00%"

    output_folder.mkdir(parents=True, exist_ok=True)
    wb.save(curve_output_path)

    # Se houver fórmulas/gráficos ligados ao arquivo, tenta recalcular e salvar
    recalculate_and_save(curve_output_path)

    return curve_output_path
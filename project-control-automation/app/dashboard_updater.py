from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def _copy_sheet_layout(source_ws, target_ws) -> None:
    """
    Copia a estrutura e o conteúdo de uma worksheet para outra.
    Isso inclui valores, fórmulas, estilos, dimensões e células mescladas.
    """

    # Copia células
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value

            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)

            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

            if source_cell.font:
                target_cell.font = copy(source_cell.font)

            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)

            if source_cell.border:
                target_cell.border = copy(source_cell.border)

            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    # Copia células mescladas
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # Copia largura das colunas
    for col_letter, dim in source_ws.column_dimensions.items():
        target_dim = target_ws.column_dimensions[col_letter]
        target_dim.width = dim.width
        target_dim.hidden = dim.hidden
        target_dim.bestFit = dim.bestFit
        target_dim.outlineLevel = dim.outlineLevel

    # Copia altura das linhas
    for row_idx, dim in source_ws.row_dimensions.items():
        target_dim = target_ws.row_dimensions[row_idx]
        target_dim.height = dim.height
        target_dim.hidden = dim.hidden
        target_dim.outlineLevel = dim.outlineLevel

    # Copia configurações gerais da aba
    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_format.defaultRowHeight = source_ws.sheet_format.defaultRowHeight
    target_ws.sheet_format.defaultColWidth = source_ws.sheet_format.defaultColWidth

    if source_ws.auto_filter and source_ws.auto_filter.ref:
        target_ws.auto_filter.ref = source_ws.auto_filter.ref

    if source_ws.print_area:
        target_ws.print_area = source_ws.print_area

    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)

    try:
        target_ws.sheet_properties.tabColor = copy(source_ws.sheet_properties.tabColor)
    except Exception:
        pass


def update_dashboard(config: dict) -> Path:
    """
    Copia a planilha inteira da LDD para a aba LD-ELFA do Dashboard.
    O arquivo final é salvo na pasta de saída.
    """
    input_folder = Path(config["paths"]["input_folder"])
    output_folder = Path(config["paths"]["output_folder"])

    ldd_file = config["paths"]["ldd_file"]
    dashboard_file = config["paths"]["dashboard_file"]

    target_sheet_name = config["dashboard"]["sheet_name"]

    ldd_path = input_folder / ldd_file
    dashboard_path = input_folder / dashboard_file

    if not ldd_path.exists():
        raise FileNotFoundError(f"LDD não encontrada: {ldd_path}")

    if not dashboard_path.exists():
        raise FileNotFoundError(f"Dashboard não encontrado: {dashboard_path}")

    # Abre os dois arquivos
    ldd_wb = load_workbook(ldd_path)
    dashboard_wb = load_workbook(dashboard_path)

    # A LDD normalmente é a primeira/única aba
    source_ws = ldd_wb.active

    # Se a aba existir, remove e recria para ficar limpa
    if target_sheet_name in dashboard_wb.sheetnames:
        old_index = dashboard_wb.sheetnames.index(target_sheet_name)
        del dashboard_wb[target_sheet_name]
        target_ws = dashboard_wb.create_sheet(title=target_sheet_name, index=old_index)
    else:
        target_ws = dashboard_wb.create_sheet(title=target_sheet_name, index=0)

    # Copia tudo da LDD para a aba do dashboard
    _copy_sheet_layout(source_ws, target_ws)

    output_folder.mkdir(parents=True, exist_ok=True)
    output_path = output_folder / dashboard_file

    if output_path.exists():
        try:
            output_path.unlink()
        except PermissionError:
            raise PermissionError(
                f"Feche o arquivo antes de executar: {output_path}"
            )

    dashboard_wb.save(output_path)
    return output_path
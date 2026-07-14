from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


DASHBOARD_PATH = Path("Lista de Documento - DashBoardC 2 1.xlsx")
CURVA_S_PATH = Path("BL1132-0001-CurvaS.xlsx")
OUTPUT_PATH = Path("output/BL1132-0001-CurvaS_atualizada.xlsx")

DASHBOARD_SHEET = "Cronograma"
DASHBOARD_CELL = "I3"

CURVA_S_SHEET = "Curva S"

LABEL_REALIZADO_ACUMULADO = "Realizado Acumulado"
LABEL_REALIZADO_SEMANAL = "Realizado Semanal"

MAX_SCAN_ROWS = 30


def last_sunday(today: date | None = None) -> date:
    if today is None:
        today = date.today()
    return today - timedelta(days=(today.weekday() + 1) % 7)


def parse_percent(value) -> float:
    """
    Converte valores como:
    - 0.23   -> 0.23
    - 23     -> 0.23
    - '23,00%' -> 0.23
    - '0,23%'   -> 0.23
    """
    if value is None:
        raise ValueError("Valor percentual vazio.")

    if isinstance(value, (int, float)):
        number = float(value)
        return number / 100 if number > 1 else number

    if isinstance(value, str):
        txt = value.strip().replace("%", "").replace(" ", "").replace(",", ".")
        number = float(txt)
        return number / 100 if number > 1 else number

    raise ValueError(f"Não foi possível interpretar o percentual: {value!r}")


def normalize_date(value) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass

    return None


def read_dashboard_real_percent() -> float:
    wb = load_workbook(DASHBOARD_PATH, data_only=True)
    ws = wb[DASHBOARD_SHEET]

    value = ws[DASHBOARD_CELL].value
    return parse_percent(value)


def find_date_header_row(ws) -> int:
    best_row = None
    best_count = 0

    for row_idx in range(1, min(ws.max_row, MAX_SCAN_ROWS) + 1):
        count = 0
        for cell in ws[row_idx]:
            if normalize_date(cell.value) is not None:
                count += 1

        if count > best_count:
            best_count = count
            best_row = row_idx

    if best_row is None or best_count < 3:
        raise ValueError("Não encontrei a linha de datas na Curva S.")

    return best_row


def build_date_column_map(ws, header_row: int) -> dict[date, int]:
    date_to_col: dict[date, int] = {}

    for cell in ws[header_row]:
        cell_date = normalize_date(cell.value)
        if cell_date is not None:
            date_to_col[cell_date] = cell.column

    if not date_to_col:
        raise ValueError("Não encontrei colunas de datas na Curva S.")

    return date_to_col


def find_row_by_label(ws, label: str) -> int:
    for row_idx in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=1).value
        if isinstance(cell_value, str) and cell_value.strip() == label:
            return row_idx

    raise ValueError(f"Não encontrei a linha com o texto: {label}")


def update_curve_s(real_percent: float) -> Path:
    wb = load_workbook(CURVA_S_PATH)
    ws = wb[CURVA_S_SHEET]

    header_row = find_date_header_row(ws)
    date_to_col = build_date_column_map(ws, header_row)

    corte = last_sunday()
    semana_anterior = corte - timedelta(days=7)

    if corte not in date_to_col:
        raise ValueError(f"Não encontrei a coluna da data de corte: {corte:%d/%m/%Y}")

    target_col = date_to_col[corte]
    previous_col = date_to_col.get(semana_anterior)

    row_acum = find_row_by_label(ws, LABEL_REALIZADO_ACUMULADO)
    row_sem = find_row_by_label(ws, LABEL_REALIZADO_SEMANAL)

    previous_acum = 0.0
    if previous_col is not None:
        prev_value = ws.cell(row=row_acum, column=previous_col).value
        if prev_value is not None and prev_value != "":
            previous_acum = parse_percent(prev_value)

    current_acum = real_percent
    current_sem = current_acum - previous_acum

    cell_acum = ws.cell(row=row_acum, column=target_col)
    cell_acum.value = current_acum
    cell_acum.number_format = "0.00%"

    cell_sem = ws.cell(row=row_sem, column=target_col)
    cell_sem.value = current_sem
    cell_sem.number_format = "0.00%"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)

    return OUTPUT_PATH


def main() -> None:
    real_percent = read_dashboard_real_percent()
    output_file = update_curve_s(real_percent)

    print(f"% Real lido do dashboard: {real_percent:.2%}")
    print(f"Curva S atualizada com sucesso: {output_file}")


if __name__ == "__main__":
    main()

#oi marcus
